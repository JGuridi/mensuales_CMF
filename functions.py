import enum
import sys
import requests
import dateutil.relativedelta
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.utils.cell import rows_from_range


def fecha():
    fecha_mes_pasado = date.today() - dateutil.relativedelta.relativedelta(months=2)    # cambiar a 1
    año, mes = fecha_mes_pasado.year, fecha_mes_pasado.month
    dia = pd.Period(fecha_mes_pasado ,freq='M').end_time.date().day
    return año, mes, dia

def mensual_balance(año, mes, id_b, banco, api_key):
    url = f'https://api.cmfchile.cl/api-sbifv3/recursos_api/balances/{año}/{mes}/instituciones/{id_b}?apikey={api_key}&formato=json'
    response = requests.get(url)
    print(f'Datos balance {banco}: {response}')
    return response.json()
 
def mensual_resultados(año, mes, id_b, banco, api_key):
    url = f'https://api.cmfchile.cl/api-sbifv3/recursos_api/resultados/{año}/{mes}/instituciones/{id_b}?apikey={api_key}&formato=json'
    response = requests.get(url)
    print(f'Datos resultado {banco}: {response}')
    return response.json()

def u_efe(año, mes, dia, api_key):
    url = f'https://api.cmfchile.cl/api-sbifv3/recursos_api/uf/{año}/{mes}/dias/{dia}?apikey={api_key}&formato=json'
    response = requests.get(url)
    # print(f'UF: {response}')
    return response.json()

def pegar(eeff, cuentas, montos, filas):
    for v,r in zip(cuentas, range(2, filas)):
        c = eeff.cell(row=r, column=1)
        c.value = v
        c.number_format = numbers.BUILTIN_FORMATS[1]

    for v,r in zip(montos, range(2, filas)):
        c = eeff.cell(row=r, column=2)
        c.value = v
        c.number_format = numbers.BUILTIN_FORMATS[1]

def sumas(archivo, sheet_name, sum):
    df = pd.read_excel(archivo, sheet_name=sheet_name)
    print(df)

    # filas ordenadas como queremos pegarlas
    orden = df.loc[:, ['Orden']]
    orden.columns = ['cuenta']
    orden = orden.astype(str)

    # filas reportadas
    todos = df.iloc[:, :2]          # sacamos las 2 primeras columnas (no incluye el 2 (3era columna))
    todos.columns = ['cuenta', 'monto']
    todos = todos.astype(str)

    # juntamos las tablas, manteniendo solo las filas de la tabla con el orden requerido
    merged = orden.merge(todos, how='left', left_on='cuenta', right_on='cuenta').dropna(how='any')
    merged['monto'] = merged['monto'].astype(float)
    merged.set_index('cuenta', inplace=True)

    # sumamos las cuentas que hay que sumar
    for m, n in enumerate(sum):
        merged.loc[f'suma {m + 1}', 'monto'] = merged.loc[n].sum()[0]
    return merged['monto'].tolist()

def pegar_2(ls, rng_p, dest):
    for idx, row in enumerate(rows_from_range(rng_p)):
        for cell in row:
            dest[cell].value = ls[idx]
            dest[cell].number_format = numbers.BUILTIN_FORMATS[1]
