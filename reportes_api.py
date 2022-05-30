import re
import time
import pandas as pd
import dateutil.relativedelta
from datetime import date
from functions import * # importamos funciones creadas
from tqdm import tqdm
from timeit import default_timer as timer
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, get_column_letter, column_index_from_string
from openpyxl.formula.translate import Translator

pd.options.mode.chained_assignment = None   # para evitar una advertencia

# -------------------------------------- api y pegado --------------------------------------

# definimos parámetros
ids = ['001', '009', '012', '014', '016', '028', '037', '039', '049', '051', '053', '055', '059', '999']
archivos = ['BancodeChile.xlsx', 'BancoInternacional.xlsx', 'BancoEstado.xlsx', 'Scotiabank.xlsx', 'BCI.xlsx', 'BICE.xlsx', 'Santander.xlsx', 'Security.xlsx', 'BancoFalabella.xlsx', 'Ripley.xlsx', 'Consorcio.xlsx', 'BTG.xlsx', 'Industria.xlsx']
api_key = '5bb0899f6fa8e7466d385a6305d93596f9df1014'

año, mes, dia = fecha() 

uf = u_efe(año, mes, dia, api_key)['UFs'][0]['Valor'] 


for j,k in zip(ids, archivos):
    pbar = tqdm(total=3, desc=f'Descargando datos para {k}', ncols=100, bar_format='{desc} {percentage:3.0f}%|{bar}|')
    # saca los datos de balance y resultados
    balance = pd.DataFrame(mensual_balance(año, mes, j, api_key)['CodigosBalances'])
    pbar.update(1)
    resultado = pd.DataFrame(mensual_resultados(año, mes, j, api_key)['CodigosEstadosDeResultado'])
    pbar.update(1)
    datos_balance = balance[['CodigoCuenta','MonedaTotal']].rename(columns={'CodigoCuenta': 'cuenta', 'MonedaTotal': j}).sort_values(by=['cuenta'])
    datos_resultados = resultado[['CodigoCuenta','MonedaTotal']].rename(columns={'CodigoCuenta': 'cuenta', 'MonedaTotal': j}).sort_values(by=['cuenta'])
    largo_datos_balance, largo_datos_resultados = len(datos_balance.cuenta), len(datos_resultados.cuenta)

    # pega los montos 
    wb, ws_balance, ws_resultado = exel(k)
    cuentas_balance, montos_balance = list(datos_balance['cuenta']), list(datos_balance[j])
    cuentas_resultado, montos_resultado = list(datos_resultados['cuenta']), list(datos_resultados[j])
    pegar(ws_balance, cuentas_balance, montos_balance, largo_datos_balance)
    pegar(ws_resultado, cuentas_resultado, montos_resultado, largo_datos_resultados)
    time.sleep(0.5)
    pbar.update(1)
    wb.save(filename=f'{k}')
    
    pbar.set_description(f'{k} actualizado en borrador')
    pbar.close()
# -------------------------------------- extensión fórmulas --------------------------------------

# sacamos las fechas de interés para reemplazar en las fórmulas
meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
month_idx = date.today().month - 3 # menos 3 porque la lista parte desde 0 (+1) y tengo que ir dos meses para atrás(+2). Dejar así porque necesitamos el mes antiguo (-3) y el "nuevo" (-2)
mes_antiguo_año = f'{meses[month_idx]}{(date.today() - dateutil.relativedelta.relativedelta(months=2)).year}'
mes_año_reemp = f'{meses[month_idx + 1]}{(date.today() - dateutil.relativedelta.relativedelta(months= 1)).year}'

pbar = tqdm(total=len(archivos), desc='Traspasando datos ordenados', ncols=100, bar_format='{desc}  {percentage:3.0f}%|{bar}|{n_fmt}/{total_fmt}')

for j in (archivos):
    # abrimos archivo
    wb = load_workbook(filename=f'{j}', keep_links=True)

    for i in ['Balance', 'Estado de Resultados', 'EERR']:
        # sacamos hoja de interés
        ws = wb[f'{i}']

        # obtenemos la última columna y su índice
        last_col = get_column_letter(ws.max_column)
        col_idx = column_index_from_string(last_col[0])
        # columna de destino
        col_dest = get_column_letter(col_idx + 1)

        for cell in ws[last_col:last_col]:
            # valor de la celda para evaluar operación
            f = cell.value
            # fila y columna de destino
            coord_string = re.split("\.|>", str(cell))[-2]
            coord = coordinate_from_string(coord_string)
            row_dest = coord[1]
            colrow_dest = f'{col_dest}{row_dest}'

            # reemplazamos las fechas que salen como referencia a otras hojas
            if mes_antiguo_año in str(f):
                f = f.replace(mes_antiguo_año, mes_año_reemp)
            
            if cell.data_type == 'f':
                ws[colrow_dest] = Translator(f, origin=coord_string).translate_formula(colrow_dest)

            elif cell.is_date:
                date = cell.value + dateutil.relativedelta.relativedelta(months=1)
                ws[colrow_dest] = date
                pass
            else: continue

    # # -------------------------------- traspaso info --------------------------------
 
    balance = wb['Balance']
    resultados = wb['Estado de Resultados']

    # sum_B1 = [['107000200', '107000300', '107000400'], ['141000200', '141000300', '141000400', '141000900'], ['150000300', '150000400', '150000500', '150000600', '150000700', '150000800'], ['170000101', '170000102'], ['190000200', '190001400', '190000800', '190000900', '190001000', '190001300', '190000300', '190000400', '190000500', '190000600', '190000700', '190001100', '190001200', '190001500', '190002900', '12500100'], ['241000200', '241000300', '241000400'], ['243000200', '243000300', '243000400'], ['244250201', '244250202', '244250203', '244250204'], ['260000300', '260000400', '260000500', '260000600', '260000900', '270000000']]
    sum_B1 = [['107000200', '107000300', '107000400'], ['141000200', '141000300', '141000400', '141000900'], ['150000300', '150000400', '150000500', '150000600', '150000700', '150000800'], ['170000101', '170000102'], ['190000200', '190001400', '190000800', '190000900', '190001000', '190001300', '190000300', '190000400', '190000500', '190000600', '190000700', '190001100', '190001200', '190001500', '190002900'], ['241000200', '241000300', '241000400'], ['243000200', '243000300', '243000400'], ['244250201', '244250202', '244250203', '244250204'], ['260000300', '260000400', '260000500', '260000600', '260000900', '270000000']]
    sum_R1 = [['442000104', '442000105'], ['442000106', '442000107', '442000108', '442000109']]

    for i in ['Balance', 'Estado de Resultados']:
        ws = wb[i]
        col = get_column_letter(ws.max_column)

        if i == 'Balance':
            datos_ordenados = sumas(j, 'B1', sum_B1)
            rng = f'{col}4:{col}{len(datos_ordenados)}'
            pegar_2(datos_ordenados, rng, balance)
        else:
            datos_ordenados = sumas(j, 'R1', sum_R1)
            rng = f'{col}4:{col}{len(datos_ordenados)}'
            pegar_2(datos_ordenados, rng, resultados)
        
    
    time.sleep(1)
    pbar.update(1)
    wb.save(filename = f'{j}')

# -------------------------------- anotamos UF --------------------------------            
pbar.set_description('Datos traspasados')
pbar.close()

wb = load_workbook(filename='Industria.xlsx')
ws = wb['EERR']
uf_row, uf_col = '4', col_dest
uf_pos = f'{uf_col}{uf_row}'
ws[uf_pos] = uf  
